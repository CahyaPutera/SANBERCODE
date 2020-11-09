# import library
import pandas as pd
import pandasql as pds
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference

# memasukkan data
penduduk = pd.read_csv('jumlah-penduduk-kota-bandung.csv').rename(columns = {'Kecamatan  ' : 'Nama Kecamatan'})
wilayah  = pd.read_csv('luas-wilayah-menurut-kecamatan-di-kota-bandung-2017.csv')

# merubah nama kecamatan yang berbeda
rep = {'Astanaanyar' : 'Astana Anyar', 'Buah Batu' : 'Buahbatu', 'Ujung Berung' : 'Ujungberung'}
penduduk['Nama Kecamatan'].replace(rep, inplace=True)

# membentuk query dengan pandasql untuk joining kedua data
query = """SELECT penduduk.[Nama Kecamatan], Jumlah_Kelurahan, Jumlah_Penduduk, wilayah.[Luas Wilayah (m2)]
FROM penduduk JOIN wilayah ON penduduk.[Nama Kecamatan] = wilayah.[Nama Kecamatan]
GROUP BY penduduk.[Nama Kecamatan]"""

# membentuk tabel KepadatanPenduduk
df = pds.sqldf(query)
df['Kepadatan Penduduk'] = df['Jumlah_Penduduk']/df['Luas Wilayah (m2)']*100

# inisiasi excel
wb = Workbook()
sheet = wb.active

# indexing 
sheet['A1'] = 'Nama Kecamatan'
sheet['B1'] = 'Kepadatan Penduduk'

# memasukkan nama kecamatan
for idx, val in enumerate(df['Nama Kecamatan'], start =2):
    i = sheet.cell(row = idx, column = 1)
    i.value = val

# memasukkan value kepadatan
for idx, val in enumerate(df['Kepadatan Penduduk'], start =2):
    i = sheet.cell(row = idx, column = 2)
    i.value = val

# membuat chart
chart = BarChart()
chart.type = 'col'
chart.style = 13
chart.title = 'Kepadatan Penduduk'
chart.y_axis.title = 'Kepadatan per 100 m2'
chart.x_axis.title = 'Kecamatan'

data = Reference(sheet, min_col=2, min_row=1, max_row=30, max_col=2)
cats = Reference(sheet, min_col=1, min_row=2, max_row=30)

chart.height = 10 
chart.width = 20 
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
sheet.add_chart(chart, 'E2')

# memberi nama file sesuai petunjuk
wb.save('Cahya.xlsx')